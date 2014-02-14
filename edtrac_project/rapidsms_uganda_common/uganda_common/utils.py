# -*- coding: utf-8 -*-
import logging
from django.conf import settings
from django.contrib.auth.models import Group
from django.core.exceptions import ValidationError
from django.db.models import Count, Sum
from django.http import HttpResponse
from eav.models import Attribute
from generic.utils import get_dates as get_dates_from_post
from poll.models import Poll, LocationResponseForm, STARTSWITH_PATTERN_TEMPLATE
from rapidsms.contrib.locations.models import Location
from rapidsms.models import Backend, Contact
from rapidsms_xforms.models import XFormSubmission, XFormSubmissionValue
from generic.sorters import SimpleSorter
from script.utils.handling import find_closest_match
import datetime
import re
from django.db.models import Max, Min
from xlrd import open_workbook
from rapidsms.models import Connection
from rapidsms_httprouter.models import Message
from django.db.models import Q
from poll.models import Response
from openpyxl.workbook import Workbook
import openpyxl
import types
from django.core.servers.basehttp import FileWrapper
from django.db import connection

logger = logging.getLogger(__name__)


def get_location_for_user(user):
    """
    if called with an argument, *user*, the location of a user returned (by district)
    """
    if user:
        try:
            l = Location.objects.get(name__iexact=user.username, type__name='district')
            logger.info("Location: %s" % l.name)
            return l
        except Location.DoesNotExist:
            try:
                if Contact.objects.filter(user=user).exclude(reporting_location=None).exists():
                    return Contact.objects.filter(user=user).exclude(reporting_location=None)[0].reporting_location
            except Exception, e:
                logger.error("Error: %s" % str(e))

    return Location.tree.root_nodes()[0]


def previous_calendar_week():
    """
    returns a datetime tuple with 2 dates: current datetime and a datetime from 7days before.
    """
    end_date = datetime.datetime.now()
    start_date = end_date - datetime.timedelta(days=7)
    return (start_date, end_date)


def previous_calendar_month():
    """
    returns a datetime tuple with 2 dates: current datetime and a datetime from 30 days before.
    """
    end_date = datetime.datetime.now()
    start_date = end_date - datetime.timedelta(days=30)
    return (start_date, end_date)


def previous_calendar_quarter():
    """
    returns a datetime tuple with 2 dates: current datetime and a datetime from 90 days before
    """
    end_date = datetime.datetime.now()
    start_date = end_date - datetime.timedelta(days=90)
    return (start_date, end_date)


TIME_RANGES = {
    'w': previous_calendar_week,
    'm': previous_calendar_month,
    'q': previous_calendar_quarter

}


def assign_backend(number):
    """
    assign a backend to a given number
    """
    country_code = getattr(settings, 'COUNTRY_CALLING_CODE', '256')
    backends = getattr(settings, 'BACKEND_PREFIXES', [('75', 'dmark'), ('71', 'dmark'), ('', 'dmark')])

    if number.startswith('0'):
        number = '%s%s' % (country_code, number[1:])
    elif number.startswith('+'):
        number = '%s' % number[1:]
    elif number[:len(country_code)] != country_code:
        number = '%s%s' % (country_code, number)
    if len(number) != 12:
        raise ValidationError("The number %s does not look valid" % number)
    try:
        int(number)
    except ValueError:
        raise ValidationError("The number %s does not look valid" % number)
    backendobj = None
    for prefix, backend in backends:
        if number[len(country_code):].startswith(prefix):
            backendobj, created = Backend.objects.get_or_create(name=backend)
            break
    return (number, backendobj)


def normalize_value(value):
    if isinstance(value, tuple(openpyxl.shared.NUMERIC_TYPES)):
        return value
    elif isinstance(value, (bool, datetime.date)):
        return value
    elif isinstance(value, types.NoneType):
        return ""
    elif isinstance(value, types.StringType):
        # print "str"+value
        return value
    elif isinstance(value, types.ListType):
        return ", ".join(value)

    elif isinstance(value, unicode):
        # print "unicode"
        # print unicodedata.normalize('NFKD', unicode(value)).encode('ascii', 'ignore')
        # openpyxl  hates unicode asciify
        return repr(value)[2:-1]

    else:
        print value
        return repr(value)


def create_workbook(data, filename, headers):
    wb = Workbook(optimized_write=True)
    ws = wb.create_sheet()
    if headers:
        ws.append(headers)

    for rowx, row in enumerate(data):
        ws.append(map(normalize_value, list(row)))

        # import pdb;pdb.set_trace()



        # for colx, value in enumerate(row):
        #   column_letter = get_column_letter((colx + 1))
        #  ws.cell('%s%s'%(column_letter, (rowx+ 1))).value = value
    # ws.auto_filter = ws.calculate_dimension()
    wb.save(filename)
    return True


class ExcelResponse(HttpResponse):
    """
    This class contains utilities that are used to produce Excel reports from datasets stored in a database or scraped
    from a form.
    """

    def __init__(self, data, output_name='excel_report.xlsx', headers=None, header=None, write_to_file=False,
                 force_csv=False):
        # Make sure we've got the right type of data to work with
        valid_data = False
        if hasattr(data, '__getitem__'):
            if isinstance(data[0], dict):
                if headers is None:
                    headers = data[0].keys()
                data = [[row[col] for col in headers] for row in data]
                #data.insert(0, headers)
            if hasattr(data[0], '__getitem__'):
                valid_data = True
        import StringIO

        output = StringIO.StringIO()
        mimetype = 'application/vnd.ms-excel'

        book_created = create_workbook(data, output_name, headers,)


        # book.save(output_name)
        # output.seek(0)
        if not write_to_file:
            super(ExcelResponse, self).__init__(FileWrapper(open(output_name)),
                                                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            self['Content-Disposition'] = 'attachment;filename="%s.%s"' % \
                                          (output_name.replace('"', '\"'), "xlsx")


def parse_district_value(value):
    """
    This function confirms whether your district does exist in a predefined list of districts.
    """
    location_template = STARTSWITH_PATTERN_TEMPLATE % '[a-zA-Z]*'
    regex = re.compile(location_template)
    toret = find_closest_match(value, Location.objects.filter(type__name='district'))
    if not toret:
        raise ValidationError(
            "We didn't recognize your district.  Please carefully type the name of your district and re-send.")
    else:
        return toret


Poll.register_poll_type('district', 'District Response', parse_district_value, db_type=Attribute.TYPE_OBJECT, \
                        view_template='polls/response_location_view.html',
                        edit_template='polls/response_location_edit.html',
                        report_columns=((('Text', 'text', True, 'message__text', SimpleSorter()), (
                            'Location', 'location', True, 'eav_values__generic_value_id', SimpleSorter()), (
                                             'Categories', 'categories', True, 'categories__category__name',
                                             SimpleSorter()))),
                        edit_form=LocationResponseForm)

GROUP_BY_WEEK = 1
GROUP_BY_MONTH = 2
GROUP_BY_DAY = 16
GROUP_BY_QUARTER = 32

months = {
    1: 'Jan',
    2: 'Feb',
    3: 'Mar',
    4: 'Apr',
    5: 'May',
    6: 'Jun',
    7: 'Jul',
    8: 'Aug',
    9: 'Sept',
    10: 'Oct',
    11: 'Nov',
    12: 'Dec'
}

quarters = {
    1: 'First',
    2: 'Second',
    3: 'Third',
    4: 'Forth'
}

GROUP_BY_SELECTS = {
    GROUP_BY_DAY: ('day', 'date(rapidsms_xforms_xformsubmission.created)',),
    GROUP_BY_WEEK: ('week', 'extract(week from rapidsms_xforms_xformsubmission.created)',),
    GROUP_BY_MONTH: ('month', 'extract(month from rapidsms_xforms_xformsubmission.created)',),
    GROUP_BY_QUARTER: ('quarter', 'extract(quarter from rapidsms_xforms_xformsubmission.created)',),
}


def total_submissions(keyword, start_date, end_date, location, extra_filters=None, group_by_timespan=None):
    """
    returns *total submission of values* from an xform; this is used to get certain values from and xform
    submitted database table.

    T%d is reverse engineered from the SQL django generates.
    hint: take a look at QuerySet.quary.alias_map in pdb
    """
    if extra_filters:
        extra_filters = dict([(str(k), v) for k, v in extra_filters.items()])
        q = XFormSubmission.objects.filter(**extra_filters)
        tnum = 8
    else:
        q = XFormSubmission.objects
        tnum = 6
    select = {
        'location_name': 'T%d.name' % tnum,
        'location_id': 'T%d.id' % tnum,
        'rght': 'T%d.rght' % tnum,
        'lft': 'T%d.lft' % tnum,
    }

    values = ['location_name', 'location_id', 'lft', 'rght']
    if group_by_timespan:
        select_value = GROUP_BY_SELECTS[group_by_timespan][0]
        select_clause = GROUP_BY_SELECTS[group_by_timespan][1]
        select.update({select_value: select_clause,
                       'year': 'extract (year from rapidsms_xforms_xformsubmission.created)', })
        values.extend([select_value, 'year'])
    if location.get_children().count() > 1:
        location_children_where = 'T%d.id in %s' % (tnum, (str(tuple(location.get_children().values_list(\
            'pk', flat=True)))))
    else:
        location_children_where = 'T%d.id = %d' % (tnum, location.get_children()[0].pk)

    return q.filter(
        xform__keyword=keyword,
        has_errors=False,
        created__lte=end_date,
        created__gte=start_date).values(
        'connection__contact__reporting_location__name').extra(
        tables=['locations_location'],
        where=[ \
            'T%d.lft <= locations_location.lft' % tnum, \
            'T%d.rght >= locations_location.rght' % tnum, \
            location_children_where]).extra(\
        select=select).values(*values).annotate(value=Count('id')).extra(order_by=['location_name'])


def total_attribute_value(attribute_slug_list, start_date, end_date, location, group_by_timespan=None):
    """
    the table name T8 is reverse engineered from the SQL django generates.
    hint: take a look at QuerySet.quary.alias_map in pdb
    """
    if type(attribute_slug_list) != list:
        attribute_slug_list = [attribute_slug_list]
    select = {
        'location_name': 'T8.name',
        'location_id': 'T8.id',
        'rght': 'T8.rght',
        'lft': 'T8.lft',
    }
    values = ['location_name', 'location_id', 'lft', 'rght']
    if group_by_timespan:
        select_value = GROUP_BY_SELECTS[group_by_timespan][0]
        select_clause = GROUP_BY_SELECTS[group_by_timespan][1]
        select.update({select_value: select_clause,
                       'year': 'extract (year from rapidsms_xforms_xformsubmission.created)', })
        values.extend([select_value, 'year'])
    if location.get_children().count() > 1:
        location_children_where = 'T8.id in %s' % (str(tuple(location.get_children().values_list(\
            'pk', flat=True))))
    else:
        location_children_where = 'T8.id = %d' % location.get_children()[0].pk
    return XFormSubmissionValue.objects.exclude(submission__connection__contact=None).filter(
        submission__has_errors=False,
        attribute__slug__in=attribute_slug_list,
        submission__created__lte=end_date,
        submission__created__gte=start_date).values(
        'submission__connection__contact__reporting_location__name').extra(
        tables=['locations_location'],
        where=[ \
            'T8.lft <= locations_location.lft',
            'T8.rght >= locations_location.rght',
            location_children_where]).extra(\
        select=select).values(*values).annotate(value=Sum('value_int')).extra(order_by=['location_name'])


def reorganize_location(key, report, report_dict):
    for rdict in report:
        location = rdict['location_id']
        report_dict.setdefault(location,
                               {'location_name': rdict['location_name'], 'diff': (rdict['rght'] - rdict['lft'])})
        report_dict[location][key] = rdict['value']


def reorganize_dictionary(key, report, report_dict, unique_key, default_values, value_key):
    for rdict in report:
        id = rdict[unique_key]
        report_dict.setdefault(id, {default_values: rdict[default_values]})
        report_dict[id][key] = rdict[value_key]


def reorganize_timespan(timespan, report, report_dict, location_list, request=None):
    for rdict in report:
        time = rdict[timespan]
        if timespan == 'month':
            time = datetime.datetime(int(rdict['year']), int(time), 1)
        elif timespan == 'week':
            time = datetime.datetime(int(rdict['year']), 1, 1) + datetime.timedelta(days=(int(time) * 7))
        elif timespan == 'quarter':
            time = datetime.datetime(int(rdict['year']), int(time) * 3, 1)

        report_dict.setdefault(time, {})
        location = rdict['location_name']
        report_dict[time][location] = rdict['value']

        if not location in location_list:
            location_list.append(location)


def get_group_by(start_date, end_date):
    """
    a function to add a group_by filter. In this case the filtering happens
    by a start_date and end_date
    """
    interval = end_date - start_date
    if interval <= datetime.timedelta(days=21):
        group_by = GROUP_BY_DAY
        prefix = 'day'
    elif datetime.timedelta(days=21) <= interval <= datetime.timedelta(days=90):
        group_by = GROUP_BY_WEEK
        prefix = 'week'
    elif datetime.timedelta(days=90) <= interval <= datetime.timedelta(days=270):
        group_by = GROUP_BY_MONTH
        prefix = 'month'
    else:
        group_by = GROUP_BY_QUARTER
        prefix = 'quarter'
    return {'group_by': group_by, 'group_by_name': prefix}


def get_xform_dates(request):
    """
    Process date variables from POST
    """
    #    dates = {}
    dates = get_dates_from_post(request)
    if ('start' in dates) and ('end' in dates):
        request.session['start_date'] = dates['start']
        request.session['end_date'] = dates['end']
    elif request.GET.get('start_date', None) and request.GET.get('end_date', None):
        request.session['start_date'] = dates['start'] = \
            datetime.datetime.fromtimestamp(int(request.GET['start_date']))
        request.session['end_date'] = dates['end'] = end_date = \
            datetime.datetime.fromtimestamp(int(request.GET['end_date']))
    elif request.session.get('start_date', None) and request.session.get('end_date', None):
        dates['start'] = request.session['start_date']
        dates['end'] = request.session['end_date']
    dts = XFormSubmission.objects.aggregate(Max('created'), Min('created'))
    dates['max'] = dts.get('created__max', None)
    dates['min'] = dts.get('created__min', None)
    return dates

def get_messages(request):
    # First we get all incoming messages for last 30 days
    # Getting all messages is so expensive
    limit_date = datetime.datetime.now() - datetime.timedelta(days=getattr(settings, 'MESSAGELOG_DAYS_LIMIT', 30))
    messages = Message.objects.filter(direction='I', date__gte=limit_date)

    # Get only messages handled by rapidsms_xforms and the polls app (this exludes opt in and opt out messages)
    messages = messages.filter(Q(application=None) | Q(application__in=['rapidsms_xforms', 'poll']))

    # Exclude XForm submissions
    messages = messages.exclude(
        pk__in=XFormSubmission.objects.exclude(message=None).\
        filter(has_errors=False, created__gte=limit_date).values_list('message__pk', flat=True))

    # Exclude Poll responses
    messages = messages.exclude(
        pk__in=Response.objects.exclude(message=None).\
        filter(has_errors=False, date__gte=limit_date).values_list('message__pk', flat=True))

    return messages

def parse_header_row(worksheet, fields):
#    fields=['telephone number','name', 'district', 'county', 'village', 'age', 'gender']
    field_cols = {}
    for col in range(worksheet.ncols):
        value = str(worksheet.cell(0, col).value).strip()
        if value.lower() in fields:
            field_cols[value.lower()] = col
    return field_cols


def parse_telephone(row, worksheet, cols):
    try:
        number = str(worksheet.cell(row, cols['telephone number']).value)
    except KeyError:
        number = str(worksheet.cell(row, cols['telephone']).value)
    return number.replace('-', '').strip().replace(' ', '')


def parse_name(row, worksheet, cols):
    try:
        name = str(worksheet.cell(row, cols['company name']).value).strip()
    except KeyError:
        name = str(worksheet.cell(row, cols['name']).value).strip()
    if name.__len__() > 0:
    #        name = str(worksheet.cell(row, cols['name']).value)
        return ' '.join([t.capitalize() for t in name.lower().split()])
    else:
        return 'Anonymous User'


def parse_district(row, worksheet, cols):
    return str(worksheet.cell(row, cols['district']).value)


def parse_village(row, worksheet, cols):
    return str(worksheet.cell(row, cols['village']).value)


def parse_birthdate(row, worksheet, cols):
    try:
        age = int(worksheet.cell(row, cols['age']).value)
        birthdate = '%d/%d/%d' % (
            datetime.datetime.now().day, datetime.datetime.now().month, datetime.datetime.now().year - age)
        return datetime.datetime.strptime(birthdate.strip(), '%d/%m/%Y')
    except ValueError:
        return None


def parse_gender(row, worksheet, cols):
    gender = str(worksheet.cell(row, cols['gender']).value)
    return gender.upper()[:1] if gender else None


def handle_excel_file(file, group, fields):
    if file:
        excel = file.read()
        workbook = open_workbook(file_contents=excel)
        worksheet = workbook.sheet_by_index(0)
        cols = parse_header_row(worksheet, fields)
        contacts = []
        duplicates = []
        invalid = []
        info = ''

        if not group:
            default_group = Group.objects.filter(name__icontains='ureporters')[0]
            group = default_group

        if worksheet.nrows > 1:
            validated_numbers = []
            for row in range(1, worksheet.nrows):
                numbers = parse_telephone(row, worksheet, cols)
                for raw_num in numbers.split('/'):
                    if raw_num[-2:] == '.0':
                        raw_num = raw_num[:-2]
                    if raw_num[:1] == '+':
                        raw_num = raw_num[1:]
                    if len(raw_num) >= 9:
                        validated_numbers.append(raw_num)
            duplicates = Connection.objects.filter(identity__in=validated_numbers).values_list('identity', flat=True)

            for row in range(1, worksheet.nrows):
                numbers = parse_telephone(row, worksheet, cols)
                if len(numbers) > 0:
                    contact = {'name': parse_name(row, worksheet, cols)}
                    district = parse_district(row, worksheet, cols) if 'district' in fields else None
                    village = parse_village(row, worksheet, cols) if 'village' in fields else None
                    birthdate = parse_birthdate(row, worksheet, cols) if 'age' in fields else None
                    gender = parse_gender(row, worksheet, cols) if 'gender' in fields else None
                    if district:
                        contact['reporting_location'] = find_closest_match(district,
                                                                           Location.objects.filter(
                                                                               kind__name='district'))
                    if village:
                        contact['village'] = find_closest_match(village, Location.objects)
                    if birthdate:
                        contact['birthdate'] = birthdate
                    if gender:
                        contact['gender'] = gender
                    if group:
                        contact['groups'] = group

                    for raw_num in numbers.split('/'):
                        if raw_num[-2:] == '.0':
                            raw_num = raw_num[:-2]
                        if raw_num[:1] == '+':
                            raw_num = raw_num[1:]
                        if len(raw_num) >= 9:
                            if raw_num not in duplicates:
                                number, backend = assign_backend(raw_num)
                                if number not in contacts and backend is not None:
                                    Connection.bulk.bulk_insert(send_pre_save=False,
                                                                identity=number,
                                                                backend=backend,
                                                                contact=contact)
                                    contacts.append(number)
                                elif backend is None:
                                    invalid.append(raw_num)

                        else:
                            invalid.append(raw_num)

            connections = Connection.bulk.bulk_insert_commit(send_post_save=False, autoclobber=True)
            contact_pks = connections.values_list('contact__pk', flat=True)

            if len(contacts) > 0:
                info = 'Contacts with numbers... ' + ' ,'.join(contacts) + " have been uploaded !\n\n"
            if len(duplicates) > 0:
                info = info + 'The following numbers already exist in the system and thus have not been uploaded: ' + ' ,'.join(
                    duplicates) + '\n\n'
            if len(invalid) > 0:
                info = info + 'The following numbers may be invalid and thus have not been added to the system: ' + ' ,'.join(
                    invalid) + '\n\n'
        else:
            info = "You seem to have uploaded an empty excel file, please fill the excel Contacts Template with contacts and upload again..."
    else:
        info = "Invalid file"
    return info


# For QOS testing
def handle_dongle_sms(message):
    if message.connection.identity in getattr(settings, 'MODEM_NUMBERS',
                                              ['256777773260', '256752145316',
                                               '256711957281', '256790403038',
                                               '256701205129']):
        Message.objects.create(direction="O", text=message.text,
                               status='Q', connection=message.connection)
        return True
    return False


def get_districts_for_user(user):
    if user:
        ret = Location.objects.filter(name__icontains=user.username, type__name='district')
        if ret:
            return ret
        else:
            return Location.objects.filter(type__name='district').order_by('name')
    return Location.objects.filter(type__name='district').order_by('name')
